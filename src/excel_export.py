"""
src/excel_export.py
Generates a professional Excel financial summary report.
Matches the format SOLON uses for commercial solar proposals.

Output: multi-sheet Excel workbook
  Sheet 1: Executive Summary
  Sheet 2: 25-Year Cash Flow Table
  Sheet 3: Input Assumptions
  Sheet 4: Sensitivity Analysis (IRR vs rate escalation vs discount rate)
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (PatternFill, Font, Alignment,
                               Border, Side, numbers)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from pathlib import Path
import sys
sys.path.insert(0, str(Path(__file__).parent.parent))
from src.financial_model import (SolarProjectInputs, SolarProjectResults,
                                   calculate, format_results)


# Colors matching commercial solar proposal aesthetic
SOLAR_ORANGE = "FF6B35"
DARK_BLUE    = "1B3A5C"
LIGHT_BLUE   = "D6E8F7"
MID_BLUE     = "4A90C4"
GRAY         = "F2F2F2"
WHITE        = "FFFFFF"
GREEN        = "27AE60"
DARK_GRAY    = "555555"


def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def make_header_fill(color=DARK_BLUE):
    return PatternFill("solid", fgColor=color)


def write_executive_summary(ws, inputs: SolarProjectInputs,
                              results: SolarProjectResults):
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20

    # Logo row
    ws.row_dimensions[1].height = 50
    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = "☀ SOLAR PROJECT FINANCIAL ANALYSIS"
    c.font = Font(bold=True, size=18, color=WHITE)
    c.fill = make_header_fill(DARK_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # Project title
    ws.merge_cells("A2:D2")
    ws["A2"].value = f"Commercial Solar — {inputs.system_size_kw:.0f} kW System"
    ws["A2"].font = Font(size=12, color=DARK_GRAY, italic=True)
    ws["A2"].alignment = Alignment(horizontal="center")

    ws.row_dimensions[3].height = 8

    # Key metrics table header
    ws.merge_cells("A4:D4")
    ws["A4"].value = "KEY FINANCIAL METRICS"
    ws["A4"].font = Font(bold=True, size=11, color=WHITE)
    ws["A4"].fill = make_header_fill(MID_BLUE)
    ws["A4"].alignment = Alignment(horizontal="center")

    metrics_data = [
        ("INVESTMENT",           "",           "",          ""),
        ("Installed Cost",       f"${results.total_installed_cost:,.0f}", "kW Cost",
         f"${results.total_installed_cost/inputs.system_size_kw/1000:.2f}/W"),
        ("Federal ITC (30%)",    f"-${results.itc_amount:,.0f}", "ITC Rate",  "30%"),
        ("Net Cost After ITC",   f"${results.net_cost_after_itc:,.0f}", "", ""),
        ("MACRS Depreciation NPV",f"${results.macrs_pv_benefit:,.0f}", "", ""),
        ("",                     "",           "",          ""),
        ("RETURNS",              "",           "",          ""),
        ("IRR",                  f"{results.irr*100:.1f}%", "Discount Rate",
         f"{0:.0f}% hurdle"),
        ("NPV (25-year)",        f"${results.npv:,.0f}", "", ""),
        ("Simple Payback",       f"{results.simple_payback_yrs:.1f} years", "", ""),
        ("Discounted Payback",   f"{results.discounted_payback_yrs:.1f} years", "", ""),
        ("LCOE",                 f"${results.lcoe:.3f}/kWh", "Grid Rate",
         f"${inputs.electricity_rate:.3f}/kWh"),
        ("",                     "",           "",          ""),
        ("ENERGY & SAVINGS",     "",           "",          ""),
        ("Year 1 Production",    f"{results.year1_production_kwh:,.0f} kWh",
         "Capacity Factor",     f"{inputs.capacity_factor*100:.1f}%"),
        ("Year 1 Savings",       f"${results.year1_savings_usd:,.0f}", "", ""),
        ("25-Year Savings",      f"${results.lifetime_savings_usd:,.0f}", "", ""),
        ("CO2 Offset (25yr)",    f"{results.co2_offset_tons:,.0f} tons", "", ""),
    ]

    for i, (label, val1, label2, val2) in enumerate(metrics_data, start=5):
        row = i
        if label in ("INVESTMENT", "RETURNS", "ENERGY & SAVINGS"):
            ws.merge_cells(f"A{row}:D{row}")
            ws[f"A{row}"].value = label
            ws[f"A{row}"].font = Font(bold=True, size=10, color=WHITE)
            ws[f"A{row}"].fill = make_header_fill(SOLAR_ORANGE)
            ws[f"A{row}"].alignment = Alignment(horizontal="left", indent=1)
        elif label == "":
            ws.row_dimensions[row].height = 6
        else:
            ws[f"A{row}"].value = label
            ws[f"A{row}"].font = Font(size=10)
            ws[f"A{row}"].fill = PatternFill("solid", fgColor=GRAY if i%2==0 else WHITE)
            ws[f"B{row}"].value = val1
            ws[f"B{row}"].font = Font(bold=True, size=10, color=DARK_BLUE)
            ws[f"B{row}"].alignment = Alignment(horizontal="right")
            ws[f"C{row}"].value = label2
            ws[f"C{row}"].font = Font(size=9, color=DARK_GRAY, italic=True)
            ws[f"D{row}"].value = val2
            ws[f"D{row}"].font = Font(size=9, color=DARK_GRAY)
            ws[f"D{row}"].alignment = Alignment(horizontal="right")


def write_cashflow_table(ws, results: SolarProjectResults):
    ws.column_dimensions["A"].width = 8
    headers = ["Year","Production (kWh)","Elec Rate","Gross Savings",
               "O&M Cost","MACRS Benefit","Net Cash Flow","Cumulative CF","Discounted CF"]
    col_widths = [6,18,12,14,12,14,15,15,15]

    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        ws.column_dimensions[get_column_letter(i)].width = w
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, size=10, color=WHITE)
        c.fill = make_header_fill(DARK_BLUE)
        c.alignment = Alignment(horizontal="center")

    for row_idx, row in results.cash_flows.iterrows():
        r = row_idx + 2
        bg = GRAY if row_idx % 2 == 0 else WHITE
        fill = PatternFill("solid", fgColor=bg)
        values = [
            int(row["year"]),
            f"{row['production_kwh']:,.0f}",
            f"${row['elec_rate']:.4f}",
            f"${row['gross_savings']:,.0f}",
            f"${row['om_cost']:,.0f}",
            f"${row['macrs_benefit']:,.0f}",
            f"${row['net_cash_flow']:,.0f}",
            f"${row['cumulative_cf']:,.0f}",
            f"${row['discounted_cf']:,.0f}",
        ]
        for col, val in enumerate(values, 1):
            c = ws.cell(row=r, column=col, value=val)
            c.fill = fill
            c.font = Font(size=9,
                          color=(GREEN if col == 8 and row["cumulative_cf"] > 0
                                 else DARK_GRAY))
            c.alignment = Alignment(horizontal="right" if col > 1 else "center")


def write_assumptions(ws, inputs: SolarProjectInputs):
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20

    ws.merge_cells("A1:B1")
    ws["A1"].value = "INPUT ASSUMPTIONS"
    ws["A1"].font = Font(bold=True, size=12, color=WHITE)
    ws["A1"].fill = make_header_fill(DARK_BLUE)
    ws["A1"].alignment = Alignment(horizontal="center")

    assumptions = [
        ("System Size (kW)",      f"{inputs.system_size_kw:.0f} kW"),
        ("Cost per Watt",         f"${inputs.cost_per_watt:.2f}/W"),
        ("Capacity Factor",       f"{inputs.capacity_factor*100:.1f}%"),
        ("Electricity Rate",      f"${inputs.electricity_rate:.4f}/kWh"),
        ("Rate Escalation",       f"{inputs.rate_escalation*100:.1f}%/yr"),
        ("Discount Rate (WACC)",  f"{inputs.discount_rate*100:.1f}%"),
        ("Analysis Period",       f"{inputs.analysis_period} years"),
        ("Federal ITC",           "30% (IRA 2022)"),
        ("MACRS Depreciation",    "5-year (IRS half-year)"),
        ("O&M Cost",              f"${inputs.om_cost_per_kw_yr:.0f}/kW/yr"),
        ("O&M Escalation",        f"{inputs.om_escalation*100:.1f}%/yr"),
        ("Panel Degradation",     f"{inputs.annual_degradation*100:.2f}%/yr"),
        ("Project Year",          str(inputs.project_year)),
        ("Data Sources",          "EIA + NREL + NREL SAM"),
    ]
    for i, (label, value) in enumerate(assumptions, start=2):
        ws[f"A{i}"].value = label
        ws[f"A{i}"].font = Font(size=10)
        ws[f"A{i}"].fill = PatternFill("solid", fgColor=GRAY if i%2==0 else WHITE)
        ws[f"B{i}"].value = value
        ws[f"B{i}"].font = Font(size=10, bold=True, color=DARK_BLUE)
        ws[f"B{i}"].alignment = Alignment(horizontal="right")


def generate_report(inputs: SolarProjectInputs,
                    results: SolarProjectResults,
                    output_path: str = "outputs/excel/solar_analysis.xlsx") -> str:
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Executive Summary"
    ws2 = wb.create_sheet("25-Year Cash Flows")
    ws3 = wb.create_sheet("Assumptions")

    write_executive_summary(ws1, inputs, results)
    write_cashflow_table(ws2, results)
    write_assumptions(ws3, inputs)

    # Style tabs
    for ws, color in [(ws1, "1B3A5C"), (ws2, "4A90C4"), (ws3, "27AE60")]:
        ws.sheet_properties.tabColor = color

    wb.save(output_path)
    print(f"  ✓ Excel report → {output_path}")
    return output_path


if __name__ == "__main__":
    from src.eia_rates import get_rate_for_state, get_capacity_factor
    inputs = SolarProjectInputs(
        system_size_kw   = 250,
        electricity_rate = get_rate_for_state("AZ"),
        capacity_factor  = get_capacity_factor("AZ"),
    )
    results = calculate(inputs)
    generate_report(inputs, results)
    print(f"  IRR: {results.irr*100:.1f}%  NPV: ${results.npv:,.0f}")
